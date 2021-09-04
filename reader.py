import os
import glob
import pickle
import hashlib
from openpyxl import load_workbook
from logger import logger


class Reader:
    def __init__(
            self,
            excel_file,
            dir_to_cv
    ):
        self.file = excel_file
        self.path_cvs = dir_to_cv
        self.cv_base = load_workbook(filename=self.file).active
        self.rows = self._get_row()
        self.current_row = int()
        self.error = None

        _create_pickle()
        self._set_cv_base_hash()
        self._get_error(self.cv_base_hash)

    def _get_row(self):
        start_row = 2
        if self.error:
            start_row = self.error

        for i in range(start_row, self.cv_base.max_row+1):
            self.current_row = i
            row = self.cv_base[i]
            yield row

    def _get_cv(self, vacancy, full_name):
        path = os.path.join(self.path_cvs, vacancy, full_name)
        files = glob.glob(f'{path}.*')
        if files:
            file_name = files[0]
            return open(file_name, 'rb')
        logger.warning(f'CV hasn\'t found')

    def _set_cv_base_hash(self):
        md5_hash = hashlib.md5()
        md5_hash.update(repr(self.cv_base).encode('utf-8'))
        self.cv_base_hash = md5_hash.hexdigest()

    def _get_error(self, file_hash=None):
        with open('data.pickle', 'rb') as error_log:
            errors = pickle.load(error_log)
            if errors:
                self.error = errors.get(file_hash)

    def set_error(self, row_number):
        with open('data.pickle', 'rb') as error_log:
            errors = pickle.load(error_log)
            errors[self.cv_base_hash] = row_number
        with open('data.pickle', 'wb') as error_log:
            pickle.dump(errors, error_log)


    @staticmethod
    def _clean(arg):
        arg = str(arg).strip()
        return arg

    def data(self):
        row = next(self.rows)
        vacancy, full_name, money, comment, raw_status = (self._clean(cell.value) for cell in row)
        file = self._get_cv(vacancy, full_name)
        if file:
            file_name = file.name.split('/')[-1]
        else:
            file_name = None
        return {
            'vacancy': vacancy,
            'full_name': full_name,
            'money': money,
            'comment': comment,
            'raw_status': raw_status,
            'file': file,
            'file_name': file_name
        }


def _create_pickle():
    if not glob.glob('data.pickle'):
        with open('data.pickle', 'wb') as f:
            pickle.dump(dict(), f)
